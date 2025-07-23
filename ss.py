from flask import Flask, render_template, request, redirect, url_for, send_from_directory, session, flash
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from functools import wraps
from flask import redirect, url_for, flash
from ldap3 import Server, Connection, ALL
import pymysql
import bcrypt

MYSQL_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'users_db',
    'port': 3306,
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

def get_db_connection():
    return pymysql.connect(**MYSQL_CONFIG)


app = Flask(__name__)
app.secret_key = 'your_secret_key'

DATABASE = 'data/feedback.db'
#USERS_FILE = 'data/users.csv'
EXCEL_FILE = 'data/staff_feedback.xlsx'

headers = [
    'ID', 'S/N', 'Activity', 'Department', 'Division', 'Start Date', 'Date of Last Update', 'Name', 'Work Done',
    'Status', 'Recommendation', 'Approval from ECOP (if any)'
]


def get_current_week_sheet_name():
    """Returns sheet name in format 'YYYY-WW' based on current date"""
    today = datetime.now()
    year, week_num, _ = today.isocalendar()
    return f"{year}-W{week_num:02d}"


def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Create users table if it doesn't exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(255) UNIQUE NOT NULL,
            password VARCHAR(255),
            email VARCHAR(255) UNIQUE,
            is_ldap TINYINT(1) DEFAULT 0
        )
    """)

    # Add is_ldap column if it doesn't exist
    cursor.execute("SHOW COLUMNS FROM users LIKE 'is_ldap'")
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE users ADD COLUMN is_ldap TINYINT(1) DEFAULT 0")

    conn.commit()
    cursor.close()
    conn.close()



def init_excel():
    if not os.path.exists('data'):
        os.makedirs('data')

    current_sheet_name = get_current_week_sheet_name()

    if not os.path.isfile(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name=current_sheet_name, index=False)

        format_sheet(current_sheet_name)
    else:
        wb = load_workbook(EXCEL_FILE)
        if current_sheet_name not in wb.sheetnames:
            wb.create_sheet(current_sheet_name)
            wb.save(EXCEL_FILE)
            format_sheet(current_sheet_name)
        wb.close()


def format_sheet(sheet_name):
    """Format a sheet with the required structure"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    ws.column_dimensions['A'].hidden = True

    for col in ['C', 'D', 'E', 'H', 'I', 'K', 'L']:
        ws.column_dimensions[col].width = 48.00

    for col in ['F', 'G', 'J']:
        ws.column_dimensions[col].width = 18.00

    header_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = header_font
        cell.fill = header_fill

    wb.save(EXCEL_FILE)


def get_next_available_row(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    wb.close()
    return row


def save_to_excel(entry):
    current_sheet_name = get_current_week_sheet_name()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[current_sheet_name]

    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1

    sn = row - 1

    def insert_line_breaks(text, max_length=60):
        if not text or not isinstance(text, str):
            return text
        words = text.split(' ')
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:
                current_line += (" " + word) if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
        return '\n'.join(lines)

    activity = insert_line_breaks(entry['Activity'])
    work_done = insert_line_breaks(entry['Work Done'])
    recommendation = insert_line_breaks(entry['Recommendation'])
    approval = insert_line_breaks(entry['Approval from ECOP (if any)'])

    ws.cell(row=row, column=1, value=entry['ID'])
    ws.cell(row=row, column=2, value=sn)
    ws.cell(row=row, column=3, value=activity)
    ws.cell(row=row, column=4, value=entry['Department'])
    ws.cell(row=row, column=5, value=entry['Division'])
    ws.cell(row=row, column=6, value=entry['Start Date'])
    ws.cell(row=row, column=7, value=None)
    ws.cell(row=row, column=8, value=entry['Name'])
    ws.cell(row=row, column=9, value=work_done)
    ws.cell(row=row, column=10, value=entry['Status'])
    ws.cell(row=row, column=11, value=recommendation)
    ws.cell(row=row, column=12, value=approval)

    for col in [3, 9, 11, 12]:
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    wb.save(EXCEL_FILE)
    wb.close()


def read_all_entries(username=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    if username:
        cursor.execute("SELECT * FROM feedback WHERE username = %s", (username,))
    else:
        cursor.execute("SELECT * FROM feedback")

    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return [{
        'ID': r['id'], 'Username': r['username'], 'Name': r['name'], 'Department': r['department'],
        'Division': r['division'], 'Activity': r['activity'], 'Work Done': r['work_done'],
        'Start Date': r['start_date'], 'Last Update': r['last_update'], 'Status': r['status'],
        'Recommendation': r['recommendation'], 'Approval from ECOP (if any)': r['ecop_approval'], 'Week': r['week']
    } for r in rows]


def update_entry(entry_id, updated_data):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query = '''
            UPDATE feedback 
            SET activity = %s, start_date = %s, last_update = %s, work_done = %s, 
                status = %s, recommendation = %s, ecop_approval = %s
            WHERE id = %s
        '''

        cursor.execute(query, (
            updated_data['Activity'], updated_data['Start Date'], updated_data['last_update'],
            updated_data['Work Done'], updated_data['Status'], updated_data['Recommendation'],
            updated_data['Approval from ECOP (if any)'], entry_id
        ))

        conn.commit()
        return True
    except Exception as e:
        print(f"Update failed: {e}")
        return False
    finally:
        cursor.close()
        conn.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_logged_in'):
            flash('Please log in to continue.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/form', methods=['GET', 'POST'])
@login_required
def form_page():
    init_excel()
    username = session['username']

    if request.method == 'POST':
        name = request.form['name']
        department = request.form['department']
        division = request.form['division']
        comment = request.form['comment']
        work_done_list = request.form.getlist('work_done[]')
        date_list = request.form.getlist('date[]')
        status_list = request.form.getlist('status[]')
        activity_list = request.form.getlist('Activity[]')
        recommendation_list = request.form.getlist('recommendation[]')

        conn = get_db_connection()
        cursor = conn.cursor()
        current_week = get_current_week_sheet_name()

        for i in range(len(work_done_list)):
            entry = {
                'Username': username,
                'Name': name,
                'Department': department,
                'Division': division,
                'Work Done': work_done_list[i],
                'Start Date': date_list[i],
                'Status': status_list[i],
                'Activity': activity_list[i],
                'Recommendation': recommendation_list[i],
                'Approval from ECOP (if any)': comment
            }

            cursor.execute('''INSERT INTO feedback 
                (username, name, department, division, activity, work_done, start_date, status, 
                 recommendation, ecop_approval, week, last_update) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                (entry['Username'], entry['Name'], entry['Department'], entry['Division'], entry['Activity'],
                 entry['Work Done'], entry['Start Date'], entry['Status'], entry['Recommendation'],
                 entry['Approval from ECOP (if any)'], current_week, None))

            entry['ID'] = cursor.lastrowid
            save_to_excel(entry)

        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('submissions'))

    return render_template('form_multi.html')



def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('is_admin'):
            flash('Admin access required.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function



@app.route('/submissions')
@login_required
def submissions():
    q = request.args.get('q', '').lower()
    is_admin = session.get('is_admin', False)
    username = session.get('username')

    entries = read_all_entries(username=None if is_admin else username)
    df = pd.DataFrame(entries)

    if q:
        df = df[df.apply(lambda row: q in str(row).lower(), axis=1)]

    return render_template('submissions.html', data=df.to_dict(orient='records'), is_admin=is_admin)



@app.route('/download')
@admin_required
def download():
    if not session.get('is_admin', False):
        return "Unauthorized", 403

    if not os.path.exists(EXCEL_FILE):
        return "Report file not found. Please submit some data first.", 404

    directory = os.path.abspath(os.path.dirname(EXCEL_FILE))
    filename = os.path.basename(EXCEL_FILE)
    current_sheet = get_current_week_sheet_name()

    return send_from_directory(
        directory=directory,
        path=filename,
        as_attachment=True,
        download_name=f"DRMD_Weekly_Report_{current_sheet}.xlsx"
    )


@app.route('/edit/<int:entry_id>', methods=['GET', 'POST'])
@login_required
def edit(entry_id):
    username = session.get('username')
    is_admin = session.get('is_admin', False)

    # Always fetch the entry directly by ID
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM feedback WHERE id = %s', (entry_id,))
    row = cursor.fetchone()
    cursor.close()
    conn.close()

    if not row:
        flash('Entry not found.', 'danger')
        return redirect(url_for('submissions'))

    # Check if user owns the entry or is admin
    if not is_admin and row['username'] != username:
        flash('Access denied: You can only edit your own entries.', 'danger')
        return redirect(url_for('submissions'))

    # Populate the entry dict (for template and update)
    entry = {
        'ID': row['id'],
        'Username': row['username'],
        'Name': row['name'],
        'Department': row['department'],
        'Division': row['division'],
        'Activity': row['activity'],
        'Work Done': row['work_done'],
        'Start Date': row['start_date'],
        'Last Update': row['last_update'],
        'Status': row['status'],
        'Recommendation': row['recommendation'],
        'Approval from ECOP (if any)': row['ecop_approval'],
        'Week': row['week']
    }

    if request.method == 'POST':
        updated_data = {
            'Activity': request.form['Activity'],
            'Start Date': request.form['date'],
            'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Work Done': request.form['work_done'],
            'Status': request.form['status'],
            'Recommendation': request.form['recommendation'],
            'Approval from ECOP (if any)': request.form['comment']
        }

        if update_entry(entry_id, updated_data):
            flash('Entry updated successfully.', 'success')
            return redirect(url_for('submissions'))
        else:
            flash('Failed to update entry.', 'danger')

    return render_template('edit.html', entry=entry)



def register_user(username, password, email):
    conn = get_db_connection()
    cursor = conn.cursor()


    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    if cursor.fetchone():
        return 'username_taken'

    cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
    if cursor.fetchone():
        return 'email_taken'

    hashed_pw = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    cursor.execute("INSERT INTO users (username, password, email) VALUES (%s, %s, %s)",
                   (username, hashed_pw, email))
    conn.commit()
    cursor.close()
    conn.close()
    return 'success'

def validate_user(username, password):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    if user and user['password'] and bcrypt.checkpw(password.encode(), user['password'].encode()):
        return True
    return False


@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        login_type = request.form.get('login_type', 'user')  # Default to user login

        if login_type == 'admin':
            # Admin login
            if username == 'admin' and password == 'adminpass123':
                session['is_admin'] = True
                session['user_logged_in'] = True
                session['username'] = 'admin'
                return redirect(url_for('submissions'))
            else:
                flash('Invalid admin credentials', 'danger')
        else:
            # User login
            if validate_user(username, password):
                session['user_logged_in'] = True
                session['username'] = username
                session['is_admin'] = False
                return redirect(url_for('form_page'))
            elif ldap_auth(username, password):
                session['user_logged_in'] = True
                session['username'] = username
                session['is_admin'] = False
                return redirect(url_for('form_page'))
            else:
                flash('Invalid username or password', 'error')

    return render_template('login.html')




def ldap_auth(username, password):
    try:
        server = Server('localhost', port=389, get_info=ALL)
        dn = f'uid={username},ou=users,dc=mycompany,dc=com'
        conn = Connection(server, dn, password, auto_bind=True)
        return True
    except Exception as e:
        print(f"LDAP auth error: {e}")
        return False

@app.route('/', methods=['GET', 'POST'])
def user_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Try CSV login
        if validate_user(username, password):
            session['user_logged_in'] = True
            session['username'] = username
            return redirect(url_for('form_page'))

        # Try LDAP login
        elif ldap_auth(username, password):
            session['user_logged_in'] = True
            session['username'] = username
            return redirect(url_for('form_page'))

        else:
            flash('Invalid username or password', 'error')

    return render_template('user_login.html')
@app.route('/reset_password', methods=['GET', 'POST'])
@login_required
def reset_password():
    username = session.get('username')

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            flash('New passwords do not match.', 'error')
            return render_template('reset_password.html')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT password, is_ldap FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()

        if user:
            if user['is_ldap']:
                flash('LDAP users cannot change passwords from this app.', 'warning')
                return redirect(url_for('form_page'))

            # Validate current password
            if bcrypt.checkpw(current_password.encode(), user['password'].encode()):
                hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
                cursor.execute("UPDATE users SET password = %s WHERE username = %s", (hashed, username))
                conn.commit()
                flash('Password updated successfully.', 'success')
            else:
                flash('Current password is incorrect.', 'error')
        else:
            flash('User not found.', 'error')

        cursor.close()
        conn.close()

    return render_template('reset_password.html',username=username)
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.route('/admin/reset_password/<username>', methods=['GET', 'POST'])
@admin_required
def admin_reset_password(username):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if user exists
        cursor.execute("SELECT password, is_ldap FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()

        if not user:
            flash("Company directory user passwords can't be changed on this app.", 'danger')
            logger.warning(f"Attempt to reset password for non-existent user: {username}")
            return redirect(url_for('submissions'))

        # Check if user is LDAP (default to False if is_ldap column is missing)
        is_ldap = user.get('is_ldap', 0) == 1
        if is_ldap:
            flash('LDAP user passwords cannot be changed from this application.', 'warning')
            logger.info(f"Blocked password reset attempt for LDAP user: {username}")
            return redirect(url_for('submissions'))

        if request.method == 'POST':
            new_password = request.form['new_password']
            confirm_password = request.form['confirm_password']

            # Validate password
            if new_password != confirm_password:
                flash('Passwords do not match.', 'error')
                return render_template('admin_reset_password.html', username=username)

            if len(new_password) <= 4 or not re.search(r'\d', new_password):
                flash('New password must be more than 4 characters and contain at least one number.', 'error')
                return render_template('admin_reset_password.html', username=username)

            # Update password
            hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
            cursor.execute("UPDATE users SET password = %s WHERE username = %s", (hashed, username))
            conn.commit()
            flash(f"Password for {username} reset successfully.", 'success')
            logger.info(f"Password reset successful for user: {username}")
            return redirect(url_for('submissions'))

        # GET request - show the password reset form
        return render_template('admin_reset_password.html', username=username)

    except Exception as e:
        flash(f"An error occurred: {str(e)}", 'danger')
        logger.error(f"Error in admin_reset_password for {username}: {str(e)}")
        return redirect(url_for('submissions'))

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()




def username_exists_in_ldap(username):
    try:
        server = Server('localhost', port=389, get_info=ALL)
        conn = Connection(
            server,
            user='cn=admin,dc=mycompany,dc=com',
            password='admin123',
            auto_bind=True
        )

        base_dn = 'ou=users,dc=mycompany,dc=com'
        search_filter = f'(uid={username})'

        if conn.search(search_base=base_dn, search_filter=search_filter, attributes=['uid']):
            print(f"LDAP search success: {conn.entries}")
            return len(conn.entries) > 0
        else:
            print(f"LDAP search failed or no entries found for {username}")
            return False
    except Exception as e:
        print(f"LDAP check failed: {e}")
        return False

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm = request.form['confirm_password']
        email = request.form['email']

        if password != confirm:
            flash('Passwords do not match.', 'error')
            return render_template('signup.html')

        if len(password) <= 4 or not re.search(r'\d', password):
            flash('Password must be more than 4 characters and contain at least one number.', 'error')
            return render_template('signup.html')

        result = register_user(username, password, email)
        if result == 'success':
            flash('Account created. Please log in.', 'success')
            return redirect(url_for('login'))  # Changed from user_login to login
        elif result == 'username_taken':
            flash('Username already taken. Try using your full name.', 'error')
        elif result == 'email_taken':
            flash('Email already used.', 'error')

    return render_template('signup.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))

if __name__ == '__main__':
    init_db()
    init_excel()
    app.run(debug=True)