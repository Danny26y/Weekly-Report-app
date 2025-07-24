import pymysql
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import bcrypt
from config import Config

def get_db_connection():
    return pymysql.connect(**Config.MYSQL_CONFIG)

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Department (
            Dept_ID INT PRIMARY KEY,
            Name VARCHAR(100) NOT NULL,
            HOD_ID INT,
            FOREIGN KEY (HOD_ID) REFERENCES users(id) ON DELETE SET NULL
        ) ENGINE=InnoDB
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Division (
            Div_ID INT PRIMARY KEY,
            Dept_ID INT NOT NULL,
            Name VARCHAR(100) NOT NULL,
            DH_ID INT,
            FOREIGN KEY (Dept_ID) REFERENCES Department(Dept_ID) ON DELETE CASCADE,
            FOREIGN KEY (DH_ID) REFERENCES users(id) ON DELETE SET NULL
        ) ENGINE=InnoDB
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(255) UNIQUE NOT NULL,
            password VARCHAR(255) DEFAULT NULL,
            is_ldap TINYINT(1) NOT NULL CHECK (is_ldap IN (0, 1)),
            Fname VARCHAR(100) NOT NULL,
            Lname VARCHAR(100) NOT NULL,
            Email VARCHAR(150) NOT NULL UNIQUE,
            Dept_ID INT NOT NULL,
            Div_ID INT NOT NULL,
            FOREIGN KEY (Dept_ID) REFERENCES Department(Dept_ID) ON DELETE RESTRICT,
            FOREIGN KEY (Div_ID) REFERENCES Division(Div_ID) ON DELETE RESTRICT
        ) ENGINE=InnoDB
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Feedback (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_ID INT NOT NULL,
            Dept_ID INT NOT NULL,
            Div_ID INT NOT NULL,
            activity TEXT,
            work_done TEXT,
            start_date DATE NOT NULL,
            status ENUM('Completed', 'Ongoing', 'Pending') NOT NULL,
            recommendation TEXT,
            ecop_approval TEXT,
            week VARCHAR(50) NOT NULL,
            last_update DATETIME,
            FOREIGN KEY (user_ID) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (Dept_ID) REFERENCES Department(Dept_ID) ON DELETE RESTRICT,
            FOREIGN KEY (Div_ID) REFERENCES Division(Div_ID) ON DELETE RESTRICT
        ) ENGINE=InnoDB
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_user_id ON Feedback(user_ID)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_feedback_week ON Feedback(week)")
    conn.commit()
    cursor.close()
    conn.close()

def init_excel():
    if not os.path.exists('data'):
        os.makedirs('data')
    current_sheet_name = get_current_week_sheet_name()
    if not os.path.isfile(Config.EXCEL_FILE):
        with pd.ExcelWriter(Config.EXCEL_FILE, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name=current_sheet_name, index=False)
        format_sheet(current_sheet_name)
    else:
        wb = load_workbook(Config.EXCEL_FILE)
        if current_sheet_name not in wb.sheetnames:
            wb.create_sheet(current_sheet_name)
            wb.save(Config.EXCEL_FILE)
            format_sheet(current_sheet_name)
        wb.close()

def get_current_week_sheet_name():
    today = datetime.now()
    year, week_num, _ = today.isocalendar()
    return f"{year}-W{week_num:02d}"

def format_sheet(sheet_name):
    wb = load_workbook(Config.EXCEL_FILE)
    ws = wb[sheet_name]
    ws.column_dimensions['A'].hidden = True
    for col in ['C', 'D', 'E', 'H', 'I', 'K', 'L']:
        ws.column_dimensions[col].width = 48.00
    for col in ['F', 'G', 'J']:
        ws.column_dimensions[col].width = 18.00
    header_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    headers = ['ID', 'S/N', 'Activity', 'Department', 'Division', 'Start Date', 'Date of Last Update',
               'Name', 'Work Done', 'Status', 'Recommendation', 'Approval from ECOP (if any)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = header_font
        cell.fill = header_fill
    wb.save(Config.EXCEL_FILE)

def get_next_available_row(sheet_name):
    wb = load_workbook(Config.EXCEL_FILE)
    ws = wb[sheet_name]
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    wb.close()
    return row

def get_department_name(dept_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT Name FROM Department WHERE Dept_ID = %s", (dept_id,))
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    return result['Name'] if result else 'Unknown'

def get_division_name(div_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT Name FROM Division WHERE Div_ID = %s", (div_id,))
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    return result['Name'] if result else 'Unknown'

def save_to_excel(entry):
    current_sheet_name = get_current_week_sheet_name()
    wb = load_workbook(Config.EXCEL_FILE)
    ws = wb[current_sheet_name]
    row = get_next_available_row(current_sheet_name)
    sn = row - 1

    def insert_line_breaks(text, max_length=60):
        if not text or not isinstance(text, str):
            return text
        words = text.split(' ')
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:
                current_line += (" " + word) if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
        return '\n'.join(lines)

    activity = insert_line_breaks(entry['Activity'])
    work_done = insert_line_breaks(entry['Work Done'])
    recommendation = insert_line_breaks(entry['Recommendation'])
    approval = insert_line_breaks(entry['Approval from ECOP (if any)'])
    dept_name = get_department_name(entry['Department'])
    div_name = get_division_name(entry['Division'])

    ws.cell(row=row, column=1, value=entry['ID'])
    ws.cell(row=row, column=2, value=sn)
    ws.cell(row=row, column=3, value=activity)
    ws.cell(row=row, column=4, value=dept_name)
    ws.cell(row=row, column=5, value=div_name)
    ws.cell(row=row, column=6, value=entry['Start Date'])
    ws.cell(row=row, column=7, value=entry.get('Last Update'))
    ws.cell(row=row, column=8, value=entry['Name'])
    ws.cell(row=row, column=9, value=work_done)
    ws.cell(row=row, column=10, value=entry['Status'])
    ws.cell(row=row, column=11, value=recommendation)
    ws.cell(row=row, column=12, value=approval)

    for col in [3, 9, 11, 12]:
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    wb.save(Config.EXCEL_FILE)
    wb.close()

def read_all_entries(username=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    if username:
        cursor.execute("""
            SELECT f.*, u.username, CONCAT(u.Fname, ' ', u.Lname) AS name, 
                   d.Name AS dept_name, dv.Name AS div_name, u.is_ldap
            FROM Feedback f 
            JOIN users u ON f.user_ID = u.id 
            JOIN Department d ON f.Dept_ID = d.Dept_ID
            JOIN Division dv ON f.Div_ID = dv.Div_ID
            WHERE u.username = %s
        """, (username,))
    else:
        cursor.execute("""
            SELECT f.*, u.username, CONCAT(u.Fname, ' ', u.Lname) AS name, 
                   d.Name AS dept_name, dv.Name AS div_name, u.is_ldap
            FROM Feedback f 
            JOIN users u ON f.user_ID = u.id
            JOIN Department d ON f.Dept_ID = d.Dept_ID
            JOIN Division dv ON f.Div_ID = dv.Div_ID
        """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [{
        'ID': r['id'], 'Username': r['username'], 'Name': r['name'],
        'Department': r['dept_name'], 'Division': r['div_name'],
        'Activity': r['activity'], 'Work Done': r['work_done'],
        'Start Date': r['start_date'], 'Last Update': r['last_update'],
        'Status': r['status'], 'Recommendation': r['recommendation'],
        'Approval from ECOP (if any)': r['ecop_approval'], 'Week': r['week'],
        'is_ldap': r['is_ldap']
    } for r in rows]

def update_entry(entry_id, updated_data):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = '''
            UPDATE Feedback 
            SET activity = %s, start_date = %s, last_update = %s, work_done = %s, 
                status = %s, recommendation = %s, ecop_approval = %s
            WHERE id = %s
        '''
        cursor.execute(query, (
            updated_data['Activity'], updated_data['Start Date'], updated_data['last_update'],
            updated_data['Work Done'], updated_data['Status'], updated_data['Recommendation'],
            updated_data['Approval from ECOP (if any)'], entry_id
        ))
        conn.commit()
        return True
    except Exception as e:
        print(f"Update failed: {e}")
        return False
    finally:
        cursor.close()
        conn.close()